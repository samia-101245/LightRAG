import openpyxl
import os
import keyboard
import pyperclip

def paste_clipboard_to_excel(output_file, sheet_name='Sheet1'):
    current_col = 1 
    current_row = 2 

    if os.path.exists(output_file):
        try:
            os.rename(output_file, output_file)
        except PermissionError:
            print(f"close {output_file} before running the script.")
            return

    def save_to_excel(data, file, sheet, row, col):
        if os.path.exists(file):
            wb = openpyxl.load_workbook(file)
        else:
            wb = openpyxl.Workbook()

        if sheet not in wb.sheetnames:
            ws = wb.create_sheet(sheet)
        else:
            ws = wb[sheet]

        ws.cell(row=row, column=col, value=data)
        wb.save(file)
        print(f"Pasted '{data}' at Column {col}, Row {row}")

    print("Copy text from an open file and first 'Ctrl+C' and then 'Ctrl + Shift + V' to paste it into Excel file")

    while True:
        keyboard.wait("ctrl+shift+v")  
        clipboard_data = pyperclip.paste().strip()

        if clipboard_data:
            lines = clipboard_data.split("\n")  

            for line in lines:
                save_to_excel(line, output_file, sheet_name, current_row, current_col)
                current_row += 1 

            print(f"Pasted {len(lines)} lines into Column {current_col}")

            if current_row > 100:  
                current_row = 1  
                current_col += 1  
if __name__ == "__main__":
    paste_clipboard_to_excel("output.xlsx", "Sheet1")